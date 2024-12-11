import openpyxl as xl
from pathlib import Path


def get_sheetnames_xlsx(filepath: str):
    """
    Get the sheetnames of an Excel file in xlsx format.

    Args:
        filepath (str): The file path to get the sheetnames.

    Returns:
        list: The list of sheetnames.
    """
    wb = xl.load_workbook(filepath, read_only=True)
    return wb.sheetnames


def is_product_file(filepath: str):
    """
    Check if the file is a product file, by checking if the sheetname contains "product". Not parsing the file.

    Args:
        filepath (str): The file path to check.

    Returns:
        bool: Whether the file is a product file or not.
    """
    if Path(filepath).suffix != ".xlsx":
        return False
    return "product" in get_sheetnames_xlsx(filepath)
